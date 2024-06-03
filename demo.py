# 导入所需的模块
import numpy as np
from math import * 
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from tkinter import *
import tkinter.font as tkFont
from tkinter import filedialog

class GUI:
    def __init__(self):
        self.root = Tk()
        
        self.fullScreenState = False                                                              
        self.root.attributes("-fullscreen", self.fullScreenState)                         # 实现全屏
        self.w, self.h = self.root.winfo_screenwidth(), self.root.winfo_screenheight()    #实现带任务栏全屏
        self.root.geometry("%dx%d" % (self.w, self.h))
        
        self.root.title('桁架FEM程序')                                                    # 窗口名
        self.ft = tkFont.Font(family='microsoft yahei', size=20, weight=tkFont.NORMAL)    #标准字体，微软雅黑
        
        self.switch = 0                              # 控制器
        
        self.interface()                                                                  #调用界面函数
        
        
    def browsefunc(self):
        self.filepath = filedialog.askopenfilename()
        self.confirmbutton = Button(self.root, 
                                    text="计算",
                                    width = 5, height =2, 
                                    command=self.calculate)
        self.confirmbutton.config(font = ("Helvetica", 40))
        self.confirmbutton.grid(row=3,column=0,columnspan=10)
    
    #主程序
    def calculate(self):
        self.switch = 1
        self.root.destroy()
        
    def interface(self):
        welcome_text = '欢迎使用桁架FEM程序！'
        guide_text = '您可以使用下方的 Browse 按钮来读入含有桁架信息的‘.xlsx’文件'
        self.Label0 = Label(self.root, 
                               justify='center',
                               anchor='center',
                               foreground = 'red', 
                               text=welcome_text, 
                               font=tkFont.Font(family='microsoft yahei', size=30, weight=tkFont.BOLD))
        self.Label1 = Label(self.root,
                               justify='center',
                               anchor='center', 
                               text=guide_text, 
                               font=self.ft)
        self.browsebutton = Button(self.root, 
                                   text="Browse",
                                   width = 10, height =5, 
                                   command=self.browsefunc)
        self.browsebutton.config(font = ("Helvetica", 40))
        self.Label0.grid(row=0,column=0,columnspan=10)
        self.Label1.grid(row=1,column=0,columnspan=10)
        self.browsebutton.grid(row=2,column=0,columnspan=10)

# 定义结点对象，包括结点坐标
class Node:
    def __init__(self, num, x, y, constraint, P_x, P_y):
        self.x = x  # 横坐标
        self.y = y  # 纵坐标
        self.num = num # 结点编号
        self.constraint = constraint #约束情况，分为'Null', 'x', 'y', 'x,y'四种可能
        self.P_x = P_x # x方向外载荷
        self.P_y = P_y # y方向外载荷
        self.num_x = 2 * num - 1
        self.num_y = 2 * num
        
        self.u_x = 0 # x方向结点位移
        self.u_y = 0 # y方向结点位移
        
    def __str__(self):
        str0 = str(self.num) + '号结点：' + str((self.x, self.y)) + '国际标准单位'
        return str0
    
    def __getitem__(self, var):
        if var == 'x':
            return self.x
        if var == 'y':
            return self.y
        if var == 'n':
            return self.num
        raise KeyError('Unknown key') 
        
    def __setitem__(self, var, value):
        if var == 'x':
            self.x = value
        elif var == 'y':
            self.y = value
        elif var == 'n':
            self.num = value
        else:    
            raise KeyError('Unknown key') 
    
    def move(self): # 根据结点位移返回新结点坐标
        self.x += self.u_x
        self.y += self.u_y
        
        
# 定义杆单元对象，由两个结点组成一根杆        
class Pole:
    def __init__(self, num, node1, node2, material, A): #依次为单元编号，第一个端点，第二个端点，单元材料，杆横截面积
        self.node1 = node1 # 第一个端点（默认为这根杆的局部坐标系原点）
        self.node2 = node2 # 第二个端点
        self.num = num # 杆编号
        self.material = material # 杆的材料，可由材料对象获取对应的刚度
        
        self.len = hypot(self.node2['x']-self.node1['x'],self.node2['y']-self.node1['y'])
        
        self.EA = self.material.E * A  # 拉压刚度 EA
        
        self.S = (self.material.E / self.len) * np.matrix([-1, 0, 1, 0]) # S = DB，称为应力矩阵,此时为全局坐标系
        
        self.ue = np.matrix([0, 0, 0, 0]).T # 初始化结点位移矩阵，在桁架计算后更新为真实位移
        
        #求转角α（弧度制）
        d_x = self.node1['x'] - self.node2['x']
        d_y = self.node1['y'] - self.node2['y']
        if d_x > 0:
            if d_y >= 0:
                self.alpha = atan(d_y / d_x)
            else:
                self.alpha = atan(d_y / d_x) + 2 * pi
        elif d_x < 0:
            self.alpha = atan(d_y / d_x) + pi
        else:
            if d_y > 0:
                self.alpha = pi / 2
            elif d_y < 0:
                self.alpha = pi * 1.5
            else:
                raise ElementError('杆单元长为0')
            
    def __str__(self):
        str0 = str(self.num) + '号杆：' + str((self.node1['n'], self.node2['n']))
        return str0
    
    def draw(self, state): # 绘制该杆件
        coordinate_x = [self.node1['x'], self.node2['x']] # 两端点的x坐标
        coordinate_y = [self.node1['y'], self.node2['y']] # 两端点的y坐标
        if state == 'old':
            plt.plot(coordinate_x, coordinate_y, color='blue') # 绘制线段
            plt.scatter(coordinate_x, coordinate_y, color='red') # 绘制端点
        if state == 'new':
            plt.plot(coordinate_x, coordinate_y, color='green') # 绘制线段
            plt.scatter(coordinate_x, coordinate_y, color='yellow') # 绘制端点
        
        
    def Ke_generator(self): # 生成该杆单元的单元刚度矩阵,返回具体矩阵形式和位移向量序号
        cosa = cos(self.alpha)
        sina = sin(self.alpha)
        # 转换矩阵T
        self.T = np.matrix([[cosa, sina, 0, 0],   
                    [-sina, cosa, 0, 0],
                    [0, 0, cosa, sina],
                    [0, 0, -sina, cosa]])
        #单元刚度矩阵Ke 
        self.Ke = (self.EA / self.len) * (self.T.T) * np.mat('1,0,-1,0; 0,0,0,0; -1,0,1,0; 0,0,0,0') * self.T
        return (self.Ke, [self.node1.num_x, self.node1.num_y, self.node2.num_x, self.node2.num_y])
    
    def solve_sigma(self):
        self.sigma = - self.S * self.T * self.ue
        return self.sigma
    
    
# 定义材料对象
class Material:
    def __init__(self, name = '默认材料', rho = 2900, E = 70e9, nu = 0.3): #默认材料参数
        self.name = name
        self.rho = rho
        self.E = E
        self.nu = nu

# 定义桁架对象        
class Truss:
    def __init__(self, nodes, poles): # 均为字典类型
        self.nodes = nodes
        self.poles = poles
        
        # 总体载荷向量
        n = 2 * len(self.nodes)
        self.f = np.matrix((
            np.zeros((n, 1))
                            ))
        for i in self.nodes:
            self.f[2*i-2, 0] = self.nodes[i].P_x
            self.f[2*i-1, 0] = self.nodes[i].P_y
            
    def show(self, state):
        if state == 'old':
            for i in self.poles:
                self.poles[i].draw('old')
        if state == 'new':
            for i in self.poles:
                self.poles[i].draw('new')

    # 组装总体刚度矩阵
    def assemble(self):
        n = 2 * len(self.nodes)
        self.K = np.matrix(
            np.zeros((n, n))
                            )# 初始化总体刚度矩阵
        Kes = []
        for i in self.poles:
            Kes.append(self.poles[i].Ke_generator()) # （单元刚度矩阵,位移向量序号列表）
        for Ke in Kes:
            nums = Ke[1] # 位移向量序号列表
            for j in range(len(nums)):
                for k in range(len(nums)):
                      self.K[nums[j]-1, nums[k]-1] += Ke[0][j, k] # 组装
        return self.K
    
    # 置1法引入位移边界条件
    def place1(self):
        for i in self.nodes:
            if self.nodes[i].constraint == 'x':
                self._renewK(i,'x')
                self.f[2*i-2, 0] = 0
            if self.nodes[i].constraint == 'y':
                self._renewK(i,'y')
                self.f[2*i-1, 0] = 0
            if self.nodes[i].constraint == 'x,y':
                self._renewK(i,'x')
                self._renewK(i,'y')
                self.f[2*i-2, 0] = 0
                self.f[2*i-1, 0] = 0
        return (self.K, self.f)
            
    
    # 置1法用到的内部方法，用于更新刚度矩阵self.K
    def _renewK(self, n, constraint):
        if constraint == 'x':
            for num in [x for x in range(2 * len(self.nodes)) if x != 2 * n - 2]:
                self.K[2*n-2, num] = 0
                self.K[num, 2*n-2] = 0
            self.K[2*n-2, 2*n-2] = 1.0
        if constraint == 'y':
            for num in [x for x in range(2 * len(self.nodes)) if x != 2 * n - 1]:
                self.K[2*n-1, num] = 0
                self.K[num, 2*n-1] = 0
            self.K[2*n-1, 2*n-1] = 1.0
    
    # 求解，返回位移向量        
    def solve(self):
        self.u = np.linalg.solve(self.K, self.f)
        
        # 画图
        self.show('old')
        for i in self.nodes:
            self.nodes[i].u_x = self.u[2*i-2, 0]
            self.nodes[i].u_y = self.u[2*i-1, 0]
            self.nodes[i].move()
            
        # 把结点位移存到杆单元实例中，用以计算杆的应力
        for j in self.poles:
            self.poles[j].ue = np.matrix([self.poles[j].node1.u_x, self.poles[j].node1.u_y, self.poles[j].node2.u_x, self.poles[j].node2.u_y]).T 
            
        self.show('new')
        
        return self.u
        
    
# 利用openpyxl库，从表格文件读取数据
def data_extract(sheet):
    rows = sheet.rows
    
    lis=[] # 用于存储每行信息的列表
    
    # 迭代读取所有的行
    for row in rows:
        row_val = [col.value for col in row]
        lis.append(row_val)
    
    # 去掉提示信息和None信息
    lis_new = lis[1:]
    for element in lis:
        if element[0] == None:
            lis_new.remove( element )
    
    return lis_new

# 主程序
if __name__ == '__main__':
    
    main_screen = GUI()
    main_screen.root.mainloop()
    while main_screen.switch == 0:
        continue
    
    folder_path = main_screen.filepath.split('/')[:-1] 
    folder_path = '/'.join(folder_path)
    filename = main_screen.filepath.split('/')[-1][:-5]
    information = load_workbook(main_screen.filepath)# 信息文件的绝对存储位置

    
    # 获取结点信息
    nodes = data_extract(information['结点信息'])
    # 获取单元信息
    poles = data_extract(information['单元信息'])
    # 获取单元材料信息
    materials = data_extract(information['单元材料信息'])
    
    # 创建各结点、杆单元、材料实例
    node_instances = {}
    pole_instances = {}
    material_instances = {}
    for i in range(len(nodes)):
        node_instances[nodes[i][0]] = Node(nodes[i][0], nodes[i][1], nodes[i][2], nodes[i][3], nodes[i][4], nodes[i][5])
        
    for k in range(len(materials)):
        material_instances[materials[k][1]] = Material(materials[k][1], materials[k][2], materials[k][3], materials[k][4])
        
    for j in range(len(poles)):
        pole_instances[poles[j][0]] = Pole(poles[j][0], node_instances[poles[j][1]], node_instances[poles[j][2]], material_instances[poles[j][3]], poles[j][4])
    
    # 创建桁架实例
    truss_instance = Truss(node_instances, pole_instances)
    
    # 建立总体刚度矩阵
    truss_instance.assemble()
    # 引入边界条件
    truss_instance.place1()
    
    result = information
    while len(result.sheetnames) > 3:
        result.remove(result.sheet[len(result.sheetnames)-1])
    place = result.create_sheet('桁架结点位移',3)
    sigma = result.create_sheet('杆单元应力',4)

    place_matrix = truss_instance.solve()

    place.append(['结点编号', 'x方向位移（单位：m）', 'y方向位移（单位：m）'])
    for i in truss_instance.nodes:
        place.append([truss_instance.nodes[i]['n'], place_matrix[2 * i - 2, 0], place_matrix[2 * i - 1, 0]])

    sigma.append(['杆单元编号','结点1编号','结点2编号','应力（单位：Pa）'])
    for j in truss_instance.poles:
        sigma.append([truss_instance.poles[j].num, 
                      truss_instance.poles[j].node1['n'], 
                      truss_instance.poles[j].node2['n'], 
                      truss_instance.poles[j].solve_sigma()[0, 0]])
    
    
    result.save(folder_path+'/'+filename+'_计算结果.xlsx')
    
    plt.axis('equal')
    plt.savefig(folder_path+'/'+filename+'.png')
    plt.show()

